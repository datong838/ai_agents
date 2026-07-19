# -*- coding: utf-8 -*-
"""
线上 Niushop → Excel：数字孪生用真实全量字段导出（非测试抽样）。

- 凭据：同目录 .env（不打印密码）
- 订单/订单行：ns_order / ns_order_goods 全部列、全部行
- 商品：ns_goods / ns_goods_sku 全部列（分 Sheet）
- Excel 单格上限约 32767：超长 HTML 截断并记 meta 表
- 默认保留真实 PII（本机孪生私有文件）；NIUSHOP_EXCEL_MASK_PII=1 时对敏感列打码
"""
from __future__ import annotations

import sys
from pathlib import Path

import pymysql
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent
ENV_PATH = ROOT / ".env"
OUT_DIR = ROOT / "fixtures" / "excel"

# Excel 单元格安全上限（留余量）
CELL_MAX = 32000

DEFAULTS = {
    "NIUSHOP_DB_HOST": "127.0.0.1",
    "NIUSHOP_DB_PORT": "13306",
    "NIUSHOP_DB_USER": "niushop",
    "NIUSHOP_DB_NAME": "niushop_b2c_v5",
    "NIUSHOP_EXCEL_MASK_PII": "0",
}

# 仅当 MASK_PII=1 时处理；列名保留，值打码
PII_COLUMNS = {
    "mobile",
    "telephone",
    "weapp_openid",
    "address",
    "full_address",
    "buyer_ip",
    "invoice_email",
    "invoice_full_address",
    "taxpayer_number",
    "longitude",
    "latitude",
}


def load_env(path: Path) -> dict[str, str]:
    data = dict(DEFAULTS)
    if not path.is_file():
        raise SystemExit(f"缺少 {path}")
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        data[k.strip()] = v.strip().strip('"').strip("'")
    if not data.get("NIUSHOP_DB_PASSWORD"):
        raise SystemExit("NIUSHOP_DB_PASSWORD 未设置")
    return data


def connect(cfg: dict[str, str], database: str | None = None):
    return pymysql.connect(
        host=cfg["NIUSHOP_DB_HOST"],
        port=int(cfg["NIUSHOP_DB_PORT"]),
        user=cfg["NIUSHOP_DB_USER"],
        password=cfg["NIUSHOP_DB_PASSWORD"],
        database=database or None,
        charset="utf8mb4",
        connect_timeout=10,
        cursorclass=pymysql.cursors.DictCursor,
    )


def table_columns(conn, table: str) -> list[str]:
    with conn.cursor() as cur:
        cur.execute(f"SHOW COLUMNS FROM `{table}`")
        return [r["Field"] for r in cur.fetchall()]


def fetch_all(conn, sql: str) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute(sql)
        return list(cur.fetchall())


def mask_phone(v: str) -> str:
    s = str(v)
    if len(s) < 7:
        return "****"
    return s[:3] + "****" + s[-4:]


def apply_pii_mask(rows: list[dict], columns: list[str]) -> list[dict]:
    out = []
    for r in rows:
        nr = dict(r)
        for c in columns:
            if c not in PII_COLUMNS or nr.get(c) in (None, ""):
                continue
            if c in ("mobile", "telephone"):
                nr[c] = mask_phone(nr[c])
            elif c in ("longitude", "latitude"):
                nr[c] = None
            else:
                nr[c] = "***"
        out.append(nr)
    return out


def sanitize_cell(val, truncations: list, table: str, col: str, row_key: str):
    if val is None:
        return None
    if isinstance(val, (bytes, bytearray)):
        val = val.decode("utf-8", errors="replace")
    if isinstance(val, str) and len(val) > CELL_MAX:
        truncations.append(f"{table}.{col}#{row_key}:len={len(val)}")
        return val[:CELL_MAX] + f"\n…[twin-excel-truncated original_len={len(val)}]"
    return val


def write_sheet(
    wb: Workbook,
    title: str,
    rows: list[dict],
    columns: list[str],
    truncations: list,
    table: str,
    pk: str,
):
    ws = wb.create_sheet(title)
    ws.append(columns)
    for r in rows:
        key = str(r.get(pk, ""))
        ws.append([sanitize_cell(r.get(c), truncations, table, c, key) for c in columns])


def export_full_table(
    conn,
    wb: Workbook,
    table: str,
    sheet: str,
    order_by: str,
    mask_pii: bool,
    truncations: list,
    pk: str,
) -> tuple[int, int]:
    cols = table_columns(conn, table)
    col_sql = ", ".join(f"`{c}`" for c in cols)
    rows = fetch_all(conn, f"SELECT {col_sql} FROM `{table}` ORDER BY {order_by}")
    if mask_pii:
        rows = apply_pii_mask(rows, cols)
    write_sheet(wb, sheet, rows, cols, truncations, table, pk)
    return len(rows), len(cols)


def main() -> int:
    cfg = load_env(ENV_PATH)
    mask_pii = cfg.get("NIUSHOP_EXCEL_MASK_PII", "0").strip() in ("1", "true", "True", "yes")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    db = cfg.get("NIUSHOP_DB_NAME") or "niushop_b2c_v5"
    print(f"连接 {cfg['NIUSHOP_DB_HOST']}:{cfg['NIUSHOP_DB_PORT']} user={cfg['NIUSHOP_DB_USER']}")
    print(f"数据库: {db} · 模式: 数字孪生全字段 · MASK_PII={int(mask_pii)}")

    truncations: list[str] = []
    conn = connect(cfg, db)
    try:
        # —— 订单孪生（全字段）——
        wb_o = Workbook()
        wb_o.remove(wb_o.active)
        n_o, c_o = export_full_table(
            conn, wb_o, "ns_order", "orders", "order_id", mask_pii, truncations, "order_id"
        )
        n_l, c_l = export_full_table(
            conn,
            wb_o,
            "ns_order_goods",
            "order_lines",
            "order_id, order_goods_id",
            mask_pii,
            truncations,
            "order_goods_id",
        )
        p_order = OUT_DIR / "mall-order.xlsx"
        wb_o.save(p_order)
        print(f"写入 {p_order.name}: orders={n_o}x{c_o}cols · lines={n_l}x{c_l}cols")

        # 兼容旧文件名：同步一份（同内容）
        p_legacy = OUT_DIR / "mall-order-sample.xlsx"
        wb_o.save(p_legacy)

        # —— 商品孪生（全字段分表）——
        wb_g = Workbook()
        wb_g.remove(wb_g.active)
        n_g, c_g = export_full_table(
            conn, wb_g, "ns_goods", "goods", "goods_id", False, truncations, "goods_id"
        )
        n_s, c_s = export_full_table(
            conn, wb_g, "ns_goods_sku", "goods_sku", "goods_id, sku_id", False, truncations, "sku_id"
        )
        # 端可见 / 分类（孪生主链相关）
        for table, sheet, order, pk in (
            ("ns_goods_weapp", "goods_weapp", "goods_id, weapp_id", "goods_id"),
            ("ns_goods_category", "goods_category", "category_id", "category_id"),
        ):
            try:
                n, c = export_full_table(
                    conn, wb_g, table, sheet, order, False, truncations, pk
                )
                print(f"  + {sheet}: {n}x{c}cols")
            except Exception as e:
                print(f"  skip {table}: {e}")
        p_goods = OUT_DIR / "mall-goods.xlsx"
        wb_g.save(p_goods)
        # 旧名对照
        wb_g.save(OUT_DIR / "mall-sku-master.xlsx")
        print(f"写入 {p_goods.name}: goods={n_g}x{c_g}cols · sku={n_s}x{c_s}cols")

        # —— CMS / 门店（全字段，文案孪生）——
        wb_c = Workbook()
        wb_c.remove(wb_c.active)
        for table, sheet, order, pk in (
            ("ns_article", "articles", "article_id", "article_id"),
            ("ns_help", "help", "id", "id"),
            ("ns_document", "documents", "id", "id"),
            ("ns_store", "stores", "store_id", "store_id"),
            ("ns_notice", "notices", "id", "id"),
        ):
            try:
                n, c = export_full_table(
                    conn, wb_c, table, sheet, order, mask_pii, truncations, pk
                )
                print(f"  catalog {sheet}: {n}x{c}cols")
            except Exception as e:
                print(f"  skip {table}: {e}")
        p_cat = OUT_DIR / "mall-catalog.xlsx"
        wb_c.save(p_cat)
        wb_c.save(OUT_DIR / "mall-catalog-index.xlsx")
        print(f"写入 {p_cat.name}")

        if truncations:
            meta = wb_o.create_sheet("_truncation_meta") if False else None
            # 单独写截断清单
            meta_path = OUT_DIR / "twin-excel-truncations.txt"
            meta_path.write_text("\n".join(truncations), encoding="utf-8")
            print(f"注意: {len(truncations)} 处单元格因 Excel 上限截断 → {meta_path.name}")
        else:
            print("无 Excel 单元格截断")
    finally:
        conn.close()
    print("完成 · 数字孪生全字段导出")
    return 0


if __name__ == "__main__":
    sys.exit(main())
