# -*- coding: utf-8 -*-
"""
线上 Niushop → Excel：数字孪生用真实全量字段导出。

表头优先使用 MySQL COLUMN COMMENT（中文名）；无 COMMENT 时回退英文字段名。
每个工作簿含 `_字段词典` Sheet：英文字段 ↔ 中文名，供 Ontology 属性中文显示映射。

凭据：同目录 .env（不打印密码）
"""
from __future__ import annotations

import sys
from pathlib import Path

import pymysql
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent
ENV_PATH = ROOT / ".env"
OUT_DIR = ROOT / "fixtures" / "excel"
CELL_MAX = 32000

DEFAULTS = {
    "NIUSHOP_DB_HOST": "127.0.0.1",
    "NIUSHOP_DB_PORT": "13306",
    "NIUSHOP_DB_USER": "niushop",
    "NIUSHOP_DB_NAME": "niushop_b2c_v5",
    "NIUSHOP_EXCEL_MASK_PII": "0",
}

PII_COLUMNS = {
    "mobile",
    "telephone",
    "weapp_openid",
    "address",
    "full_address",
    "buyer_ip",
    "invoice_email",
    "invoice_full_address",
    "taxpayer_number",
    "longitude",
    "latitude",
}


def load_env(path: Path) -> dict[str, str]:
    data = dict(DEFAULTS)
    if not path.is_file():
        raise SystemExit(f"缺少 {path}")
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        data[k.strip()] = v.strip().strip('"').strip("'")
    if not data.get("NIUSHOP_DB_PASSWORD"):
        raise SystemExit("NIUSHOP_DB_PASSWORD 未设置")
    return data


def connect(cfg: dict[str, str], database: str | None = None):
    return pymysql.connect(
        host=cfg["NIUSHOP_DB_HOST"],
        port=int(cfg["NIUSHOP_DB_PORT"]),
        user=cfg["NIUSHOP_DB_USER"],
        password=cfg["NIUSHOP_DB_PASSWORD"],
        database=database or None,
        charset="utf8mb4",
        connect_timeout=10,
        cursorclass=pymysql.cursors.DictCursor,
    )


def table_meta(conn, table: str) -> list[dict]:
    """返回 [{field, type, comment, zh_name}, ...]"""
    with conn.cursor() as cur:
        cur.execute(f"SHOW FULL COLUMNS FROM `{table}`")
        rows = cur.fetchall()
    out = []
    for r in rows:
        field = r["Field"]
        comment = (r.get("Comment") or "").strip()
        zh = comment if comment else field
        out.append(
            {
                "field": field,
                "type": r.get("Type") or "",
                "comment": comment,
                "zh_name": zh,
            }
        )
    return out


def fetch_all(conn, sql: str) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute(sql)
        return list(cur.fetchall())


def mask_phone(v: str) -> str:
    s = str(v)
    if len(s) < 7:
        return "****"
    return s[:3] + "****" + s[-4:]


def apply_pii_mask(rows: list[dict], fields: list[str]) -> list[dict]:
    out = []
    for r in rows:
        nr = dict(r)
        for c in fields:
            if c not in PII_COLUMNS or nr.get(c) in (None, ""):
                continue
            if c in ("mobile", "telephone"):
                nr[c] = mask_phone(nr[c])
            elif c in ("longitude", "latitude"):
                nr[c] = None
            else:
                nr[c] = "***"
        out.append(nr)
    return out


def sanitize_cell(val, truncations: list, table: str, col: str, row_key: str):
    if val is None:
        return None
    if isinstance(val, (bytes, bytearray)):
        val = val.decode("utf-8", errors="replace")
    if isinstance(val, str) and len(val) > CELL_MAX:
        truncations.append(f"{table}.{col}#{row_key}:len={len(val)}")
        return val[:CELL_MAX] + f"\n…[twin-excel-truncated original_len={len(val)}]"
    return val


def write_data_sheet(
    wb: Workbook,
    sheet_title: str,
    meta: list[dict],
    rows: list[dict],
    truncations: list,
    table: str,
    pk: str,
):
    """第1行：中文名；第2行：英文字段名（孪生映射键）；第3行起：数据。"""
    ws = wb.create_sheet(sheet_title)
    fields = [m["field"] for m in meta]
    zh_headers = [m["zh_name"] for m in meta]
    ws.append(zh_headers)
    ws.append(fields)
    for r in rows:
        key = str(r.get(pk, ""))
        ws.append(
            [sanitize_cell(r.get(c), truncations, table, c, key) for c in fields]
        )


def append_dict_sheet(wb: Workbook, table_sheets: list[tuple[str, str, list[dict]]]):
    """_字段词典：表名 / Sheet / 英文字段 / 中文名 / 类型 / COMMENT原文"""
    ws = wb.create_sheet("_字段词典", 0)
    ws.append(["物理表", "Sheet", "英文字段", "中文名", "类型", "COMMENT原文", "无中文COMMENT"])
    for table, sheet, meta in table_sheets:
        for m in meta:
            missing = "Y" if not m["comment"] else ""
            ws.append(
                [
                    table,
                    sheet,
                    m["field"],
                    m["zh_name"],
                    m["type"],
                    m["comment"],
                    missing,
                ]
            )


def export_full_table(
    conn,
    wb: Workbook,
    table: str,
    sheet: str,
    order_by: str,
    mask_pii: bool,
    truncations: list,
    pk: str,
) -> tuple[int, int, list[dict]]:
    meta = table_meta(conn, table)
    fields = [m["field"] for m in meta]
    col_sql = ", ".join(f"`{c}`" for c in fields)
    rows = fetch_all(conn, f"SELECT {col_sql} FROM `{table}` ORDER BY {order_by}")
    if mask_pii:
        rows = apply_pii_mask(rows, fields)
    write_data_sheet(wb, sheet, meta, rows, truncations, table, pk)
    return len(rows), len(fields), meta


def build_workbook(
    conn,
    specs: list[tuple[str, str, str, str]],
    mask_pii: bool,
    truncations: list,
) -> tuple[Workbook, list[str]]:
    """specs: (table, sheet, order_by, pk)"""
    wb = Workbook()
    wb.remove(wb.active)
    dict_entries: list[tuple[str, str, list[dict]]] = []
    summaries: list[str] = []
    for table, sheet, order_by, pk in specs:
        try:
            n, c, meta = export_full_table(
                conn, wb, table, sheet, order_by, mask_pii, truncations, pk
            )
            dict_entries.append((table, sheet, meta))
            no_zh = sum(1 for m in meta if not m["comment"])
            summaries.append(f"{sheet}: {n}行×{c}列(无COMMENT={no_zh})")
        except Exception as e:
            summaries.append(f"{sheet}: SKIP {e}")
    if dict_entries:
        append_dict_sheet(wb, dict_entries)
    return wb, summaries


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass

    cfg = load_env(ENV_PATH)
    mask_pii = cfg.get("NIUSHOP_EXCEL_MASK_PII", "0").strip() in ("1", "true", "True", "yes")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    db = cfg.get("NIUSHOP_DB_NAME") or "niushop_b2c_v5"
    print(f"连接 {cfg['NIUSHOP_DB_HOST']}:{cfg['NIUSHOP_DB_PORT']} user={cfg['NIUSHOP_DB_USER']}")
    print(f"数据库: {db} · 全字段+中文表头 · MASK_PII={int(mask_pii)}")

    truncations: list[str] = []
    conn = connect(cfg, db)
    try:
        # 订单
        wb_o, sums = build_workbook(
            conn,
            [
                ("ns_order", "订单", "order_id", "order_id"),
                ("ns_order_goods", "订单行", "order_id, order_goods_id", "order_goods_id"),
            ],
            mask_pii,
            truncations,
        )
        p_order = OUT_DIR / "mall-order.xlsx"
        wb_o.save(p_order)
        wb_o.save(OUT_DIR / "mall-order-sample.xlsx")
        print(f"写入 {p_order.name}: " + " · ".join(sums))

        # 商品
        wb_g, sums = build_workbook(
            conn,
            [
                ("ns_goods", "商品", "goods_id", "goods_id"),
                ("ns_goods_sku", "商品SKU", "goods_id, sku_id", "sku_id"),
                ("ns_goods_weapp", "商品端可见", "goods_id, weapp_id", "goods_id"),
                ("ns_goods_category", "商品分类", "category_id", "category_id"),
            ],
            False,
            truncations,
        )
        p_goods = OUT_DIR / "mall-goods.xlsx"
        wb_g.save(p_goods)
        wb_g.save(OUT_DIR / "mall-sku-master.xlsx")
        print(f"写入 {p_goods.name}: " + " · ".join(sums))

        # 内容/门店
        wb_c, sums = build_workbook(
            conn,
            [
                ("ns_article", "文章", "article_id", "article_id"),
                ("ns_help", "帮助", "id", "id"),
                ("ns_document", "协议文档", "id", "id"),
                ("ns_store", "门店", "store_id", "store_id"),
                ("ns_notice", "公告", "id", "id"),
            ],
            mask_pii,
            truncations,
        )
        p_cat = OUT_DIR / "mall-catalog.xlsx"
        wb_c.save(p_cat)
        wb_c.save(OUT_DIR / "mall-catalog-index.xlsx")
        print(f"写入 {p_cat.name}: " + " · ".join(sums))

        if truncations:
            meta_path = OUT_DIR / "twin-excel-truncations.txt"
            meta_path.write_text("\n".join(truncations), encoding="utf-8")
            print(f"注意: {len(truncations)} 处截断 → {meta_path.name}")
        else:
            print("无 Excel 单元格截断")
    finally:
        conn.close()
    print("完成 · 中文表头孪生导出")
    return 0


if __name__ == "__main__":
    sys.exit(main())
